"""
FamilyHub Development Tracker - Export Project Data

Management command to export project data to various formats for backup and reporting.
Supports JSON, CSV, and Excel export formats.

Usage:
    python manage.py export_project_data --format json
    python manage.py export_project_data --format csv --project 1
    python manage.py export_project_data --format excel --output /path/to/export.xlsx
"""
from django.core.management.base import BaseCommand, CommandError
from django.core.serializers import serialize
from django.http import HttpResponse
from django.utils import timezone
import json
import csv
import os
from datetime import datetime

from tracker.models import Project, Application, Artifact, Task, Decision, Integration


class Command(BaseCommand):
    help = 'Export project data to JSON, CSV, or Excel format'

    def add_arguments(self, parser):
        parser.add_argument(
            '--format',
            type=str,
            choices=['json', 'csv', 'excel'],
            default='json',
            help='Export format (default: json)',
        )
        parser.add_argument(
            '--project',
            type=int,
            help='Export specific project ID (default: all projects)',
        )
        parser.add_argument(
            '--output',
            type=str,
            help='Output file path (default: exports/familyhub_export_YYYYMMDD)',
        )
        parser.add_argument(
            '--include',
            type=str,
            nargs='+',
            choices=['projects', 'applications', 'tasks', 'artifacts', 'decisions', 'integrations'],
            default=['projects', 'applications', 'tasks', 'artifacts', 'decisions', 'integrations'],
            help='Include specific data types (default: all)',
        )

    def handle(self, *args, **options):
        try:
            # Validate project ID if provided
            project = None
            if options['project']:
                try:
                    project = Project.objects.get(pk=options['project'])
                    self.stdout.write(f'Exporting data for project: {project.name}')
                except Project.DoesNotExist:
                    raise CommandError(f'Project with ID {options["project"]} not found')
            else:
                self.stdout.write('Exporting data for all projects')

            # Create output directory if needed
            output_path = self.get_output_path(options['output'], options['format'])
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Export data based on format
            export_format = options['format']
            include_types = options['include']

            if export_format == 'json':
                self.export_json(output_path, project, include_types)
            elif export_format == 'csv':
                self.export_csv(output_path, project, include_types)
            elif export_format == 'excel':
                self.export_excel(output_path, project, include_types)

            self.stdout.write(
                self.style.SUCCESS(f'Successfully exported data to: {output_path}')
            )

        except Exception as e:
            raise CommandError(f'Error exporting data: {str(e)}')

    def get_output_path(self, custom_path, export_format):
        """Generate output file path."""
        if custom_path:
            return custom_path

        # Create default path
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'familyhub_export_{timestamp}.{export_format}'
        
        # Create exports directory in project root
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
        exports_dir = os.path.join(base_dir, 'exports')
        
        return os.path.join(exports_dir, filename)

    def get_filtered_data(self, project, include_types):
        """Get filtered data based on project and include types."""
        data = {}

        if 'projects' in include_types:
            if project:
                data['projects'] = [project]
            else:
                data['projects'] = list(Project.objects.all())

        if 'applications' in include_types:
            if project:
                data['applications'] = list(project.application_set.all())
            else:
                data['applications'] = list(Application.objects.all())

        if 'tasks' in include_types:
            if project:
                data['tasks'] = list(Task.objects.filter(project=project))
            else:
                data['tasks'] = list(Task.objects.all())

        if 'artifacts' in include_types:
            if project:
                data['artifacts'] = list(Artifact.objects.filter(application__project=project))
            else:
                data['artifacts'] = list(Artifact.objects.all())

        if 'decisions' in include_types:
            if project:
                data['decisions'] = list(project.decision_set.all())
            else:
                data['decisions'] = list(Decision.objects.all())

        if 'integrations' in include_types:
            if project:
                data['integrations'] = list(Integration.objects.filter(
                    from_application__project=project
                ))
            else:
                data['integrations'] = list(Integration.objects.all())

        return data

    def export_json(self, output_path, project, include_types):
        """Export data to JSON format."""
        self.stdout.write('Exporting to JSON format...')
        
        data = self.get_filtered_data(project, include_types)
        export_data = {
            'export_info': {
                'timestamp': timezone.now().isoformat(),
                'format': 'json',
                'project': project.name if project else 'all',
                'include_types': include_types,
            },
            'data': {}
        }

        # Serialize each data type
        for data_type, objects in data.items():
            if objects:
                # Convert Django model instances to serializable format
                serialized_objects = []
                for obj in objects:
                    obj_data = {
                        'id': obj.id,
                        'created_at': obj.created_at.isoformat() if hasattr(obj, 'created_at') else None,
                        'updated_at': obj.updated_at.isoformat() if hasattr(obj, 'updated_at') else None,
                    }

                    # Add model-specific fields
                    if hasattr(obj, 'name'):
                        obj_data['name'] = obj.name
                    if hasattr(obj, 'title'):
                        obj_data['title'] = obj.title
                    if hasattr(obj, 'description'):
                        obj_data['description'] = obj.description
                    if hasattr(obj, 'status'):
                        obj_data['status'] = obj.status

                    # Add relationship fields
                    if hasattr(obj, 'project'):
                        obj_data['project_id'] = obj.project.id if obj.project else None
                        obj_data['project_name'] = obj.project.name if obj.project else None
                    if hasattr(obj, 'application'):
                        obj_data['application_id'] = obj.application.id if obj.application else None
                        obj_data['application_name'] = obj.application.name if obj.application else None

                    # Add model-specific fields
                    if data_type == 'projects':
                        obj_data.update({
                            'owner': obj.owner.username,
                            'start_date': obj.start_date.isoformat() if obj.start_date else None,
                            'target_date': obj.target_date.isoformat() if obj.target_date else None,
                        })
                    elif data_type == 'applications':
                        obj_data.update({
                            'features': obj.features,
                            'tech_stack': obj.tech_stack,
                            'version': obj.version,
                            'repository_url': obj.repository_url,
                        })
                    elif data_type == 'tasks':
                        obj_data.update({
                            'priority': obj.priority,
                            'due_date': obj.due_date.isoformat() if obj.due_date else None,
                            'assigned_to': obj.assigned_to.username if hasattr(obj, 'assigned_to') and obj.assigned_to else None,
                        })
                    elif data_type == 'artifacts':
                        obj_data.update({
                            'artifact_type': obj.artifact_type,
                            'version': obj.version,
                            'file_size': obj.file_size if hasattr(obj, 'file_size') else None,
                        })

                    serialized_objects.append(obj_data)

                export_data['data'][data_type] = serialized_objects

        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

        self.stdout.write(f'Exported {sum(len(v) for v in export_data["data"].values())} records')

    def export_csv(self, output_path, project, include_types):
        """Export data to CSV format (creates separate CSV for each data type)."""
        self.stdout.write('Exporting to CSV format...')
        
        data = self.get_filtered_data(project, include_types)
        
        # Create CSV files for each data type
        base_path = output_path.rsplit('.', 1)[0]
        
        for data_type, objects in data.items():
            if not objects:
                continue
                
            csv_path = f'{base_path}_{data_type}.csv'
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                if data_type == 'projects':
                    fieldnames = ['id', 'name', 'description', 'status', 'owner', 'start_date', 'target_date', 'created_at', 'updated_at']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for obj in objects:
                        writer.writerow({
                            'id': obj.id,
                            'name': obj.name,
                            'description': obj.description,
                            'status': obj.status,
                            'owner': obj.owner.username,
                            'start_date': obj.start_date,
                            'target_date': obj.target_date,
                            'created_at': obj.created_at,
                            'updated_at': obj.updated_at,
                        })

                elif data_type == 'applications':
                    fieldnames = ['id', 'name', 'description', 'status', 'project', 'version', 'repository_url', 'created_at', 'updated_at']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for obj in objects:
                        writer.writerow({
                            'id': obj.id,
                            'name': obj.name,
                            'description': obj.description,
                            'status': obj.status,
                            'project': obj.project.name,
                            'version': obj.version,
                            'repository_url': obj.repository_url,
                            'created_at': obj.created_at,
                            'updated_at': obj.updated_at,
                        })

                elif data_type == 'tasks':
                    fieldnames = ['id', 'title', 'description', 'status', 'priority', 'project', 'application', 'due_date', 'created_at', 'updated_at']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for obj in objects:
                        writer.writerow({
                            'id': obj.id,
                            'title': obj.title,
                            'description': obj.description,
                            'status': obj.status,
                            'priority': obj.priority,
                            'project': obj.project.name if obj.project else '',
                            'application': obj.application.name if obj.application else '',
                            'due_date': obj.due_date,
                            'created_at': obj.created_at,
                            'updated_at': obj.updated_at,
                        })

            self.stdout.write(f'Created CSV: {csv_path} ({len(objects)} records)')

    def export_excel(self, output_path, project, include_types):
        """Export data to Excel format with multiple sheets."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            import pandas as pd
        except ImportError:
            raise CommandError('openpyxl and pandas are required for Excel export. Install with: pip install openpyxl pandas')

        self.stdout.write('Exporting to Excel format...')
        
        data = self.get_filtered_data(project, include_types)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet if it exists
        if wb.active:
            wb.remove(wb.active)
        
        for data_type, objects in data.items():
            if not objects:
                continue
                
            # Create worksheet for this data type
            ws = wb.create_sheet(title=data_type.capitalize())
            
            # Convert objects to DataFrame
            if data_type == 'projects':
                df_data = []
                for obj in objects:
                    df_data.append({
                        'ID': obj.id,
                        'Name': obj.name,
                        'Description': obj.description,
                        'Status': obj.status,
                        'Owner': obj.owner.username,
                        'Start Date': obj.start_date,
                        'Target Date': obj.target_date,
                        'Created': obj.created_at,
                        'Updated': obj.updated_at,
                    })
            
            elif data_type == 'applications':
                df_data = []
                for obj in objects:
                    df_data.append({
                        'ID': obj.id,
                        'Name': obj.name,
                        'Description': obj.description,
                        'Status': obj.status,
                        'Project': obj.project.name,
                        'Version': obj.version,
                        'Repository': obj.repository_url,
                        'Created': obj.created_at,
                        'Updated': obj.updated_at,
                    })
            
            elif data_type == 'tasks':
                df_data = []
                for obj in objects:
                    df_data.append({
                        'ID': obj.id,
                        'Title': obj.title,
                        'Description': obj.description,
                        'Status': obj.status,
                        'Priority': obj.priority,
                        'Project': obj.project.name if obj.project else '',
                        'Application': obj.application.name if obj.application else '',
                        'Due Date': obj.due_date,
                        'Created': obj.created_at,
                        'Updated': obj.updated_at,
                    })
            
            # Create DataFrame and add to worksheet
            if df_data:
                df = pd.DataFrame(df_data)
                
                # Add DataFrame to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # Format headers
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add summary sheet
        summary_ws = wb.create_sheet(title='Summary', index=0)
        summary_ws.append(['FamilyHub Development Tracker - Export Summary'])
        summary_ws.append(['Export Date:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_ws.append(['Project:', project.name if project else 'All Projects'])
        summary_ws.append([])
        summary_ws.append(['Data Type', 'Record Count'])
        
        for data_type, objects in data.items():
            summary_ws.append([data_type.capitalize(), len(objects)])
        
        # Format summary sheet
        summary_ws['A1'].font = Font(bold=True, size=14)
        
        # Save workbook
        wb.save(output_path)
        
        total_records = sum(len(objects) for objects in data.values())
        self.stdout.write(f'Created Excel file with {len(data)} sheets and {total_records} total records')
